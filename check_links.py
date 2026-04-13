import argparse
import concurrent.futures
import configparser
import json
import threading
import time
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Callable, Optional
from urllib.parse import urlparse

import openpyxl
import requests
from openpyxl.styles import PatternFill


STATUS_OK = "ok"
STATUS_REDIRECT = "redirect"
STATUS_BLACKLIST = "blacklist"
STATUS_EMPTY = "empty"
STATUS_INVALID = "invalid"
STATUS_BLOCKED = "blocked"
STATUS_FAILED = "failed"

TECH_OK = "ok"
TECH_REDIRECT = "redirect"
TECH_BLACKLIST = "blacklist"
TECH_EMPTY = "empty"
TECH_INVALID = "invalid_format"
TECH_TIMEOUT = "timeout"
TECH_REQUEST_ERROR = "request_error"
TECH_BOT_CHECK = "bot_check"
TECH_DELETED = "deleted_page"

PROBLEM_STATUSES = {
    STATUS_BLACKLIST,
    STATUS_EMPTY,
    STATUS_INVALID,
    STATUS_BLOCKED,
    STATUS_FAILED,
}

STATUS_COLORS = {
    STATUS_BLACKLIST: "F4CCCC",
    STATUS_EMPTY: "EAD1DC",
    STATUS_INVALID: "F4CCCC",
    STATUS_BLOCKED: "FCE5CD",
    STATUS_FAILED: "FCE5CD",
}

AUTO_COLUMN_CANDIDATES = {
    "url",
    "link",
    "ссылка",
    "url статьи",
    "ссылка на статью",
    "article url",
    "article link",
}

DEFAULT_SETTINGS = {
    "input_file": None,
    "all_sheets": False,
    "sheet": None,
    "column_name": None,
    "column_index": None,
    "blacklist_file": "blacklist.txt",
    "output_file": None,
    "timeout": 8,
    "workers": 12,
    "details": False,
}

DEFAULT_KNOWN_NAMES = {
    "known_columns": [],
    "file_rules": [],
}


def get_run_date_label() -> str:
    return datetime.now().strftime("%d-%m %H-%M")


def normalize_url(raw_value: Optional[str]) -> str:
    if raw_value is None:
        return ""

    value = str(raw_value).strip()
    if not value:
        return ""

    lowered = value.lower()

    if lowered.startswith(("mailto:", "javascript:", "tel:")):
        return value

    if not lowered.startswith(("http://", "https://")):
        value = "https://" + value

    return value


def extract_domain(url: str) -> str:
    try:
        parsed = urlparse(url)
        domain = parsed.netloc.lower().strip()

        if domain.startswith("www."):
            domain = domain[4:]

        return domain
    except Exception:
        return ""


def is_valid_url(url: str) -> bool:
    if not url:
        return False

    lowered = url.lower()

    if lowered.startswith(("mailto:", "javascript:", "tel:")):
        return False

    try:
        parsed = urlparse(url)
        return parsed.scheme in ("http", "https") and bool(parsed.netloc)
    except Exception:
        return False


def load_blacklist(file_path: str) -> set[str]:
    path = Path(file_path)

    if not path.exists():
        return set()

    blacklist = set()

    with path.open("r", encoding="utf-8") as file:
        for line in file:
            domain = line.strip().lower()

            if not domain or domain.startswith("#"):
                continue

            if domain.startswith("www."):
                domain = domain[4:]

            blacklist.add(domain)

    return blacklist


def domain_in_blacklist(domain: str, blacklist: set[str]) -> bool:
    if not domain:
        return False

    if domain in blacklist:
        return True

    return any(domain.endswith("." + blocked) for blocked in blacklist)


def build_technical_status(response: requests.Response) -> str:
    if response.ok:
        if response.history:
            return TECH_REDIRECT
        return TECH_OK

    return f"http_{response.status_code}"


def looks_like_bot_check(response: requests.Response) -> bool:
    indicators = [
        "captcha",
        "cloudflare",
        "verify you are human",
        "verify you are a human",
        "are you human",
        "bot check",
        "access denied",
        "ddos protection",
        "attention required",
    ]

    url_text = str(response.url).lower()

    for indicator in indicators:
        if indicator in url_text:
            return True

    server_header = response.headers.get("Server", "").lower()
    if response.status_code in (403, 429) and "cloudflare" in server_header:
        return True

    content_type = response.headers.get("Content-Type", "").lower()
    if "text/html" in content_type:
        try:
            body = response.text[:5000].lower()
            for indicator in indicators:
                if indicator in body:
                    return True
        except Exception:
            return False

    return False


def looks_like_deleted_page(response: requests.Response) -> bool:
    if response.status_code in (404, 410):
        return True

    content_type = response.headers.get("Content-Type", "").lower()
    if "text/html" not in content_type:
        return False

    markers = [
        "страница не найдена",
        "материал не найден",
        "новость не найдена",
        "публикация не найдена",
        "страница удалена",
        "материал удален",
        "page not found",
        "article not found",
        "story not found",
        "content not found",
        "this page is no longer available",
        "the page you requested could not be found",
        "410 gone",
        "404 not found",
    ]

    try:
        body = response.text[:7000].lower()
    except Exception:
        return False

    return any(marker in body for marker in markers)


def _perform_request(url: str, headers: dict, timeout: int) -> requests.Response:
    return requests.get(
        url,
        headers=headers,
        timeout=timeout,
        allow_redirects=True,
    )


def check_one_url(raw_url: Optional[str], blacklist: set[str], timeout: int) -> tuple[str, str]:
    normalized_url = normalize_url(raw_url)

    if not normalized_url:
        return STATUS_EMPTY, TECH_EMPTY

    if not is_valid_url(normalized_url):
        return STATUS_INVALID, TECH_INVALID

    domain = extract_domain(normalized_url)

    if domain_in_blacklist(domain, blacklist):
        return STATUS_BLACKLIST, TECH_BLACKLIST

    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/122.0.0.0 Safari/537.36"
        )
    }

    retry_timeout = min(timeout * 2, timeout + 10)

    try:
        response = _perform_request(normalized_url, headers, timeout)
    except requests.exceptions.Timeout:
        try:
            time.sleep(0.4)
            response = _perform_request(normalized_url, headers, retry_timeout)
        except requests.exceptions.Timeout:
            return STATUS_FAILED, TECH_TIMEOUT
        except requests.exceptions.RequestException:
            return STATUS_FAILED, TECH_REQUEST_ERROR
    except requests.exceptions.RequestException:
        try:
            time.sleep(0.4)
            response = _perform_request(normalized_url, headers, retry_timeout)
        except requests.exceptions.Timeout:
            return STATUS_FAILED, TECH_TIMEOUT
        except requests.exceptions.RequestException:
            return STATUS_FAILED, TECH_REQUEST_ERROR

    technical_status = build_technical_status(response)

    if looks_like_deleted_page(response):
        return STATUS_FAILED, TECH_DELETED

    if looks_like_bot_check(response):
        if response.status_code in (403, 429):
            return STATUS_BLOCKED, TECH_BOT_CHECK
        return STATUS_FAILED, TECH_BOT_CHECK

    if response.ok:
        if response.history:
            return STATUS_REDIRECT, technical_status
        return STATUS_OK, technical_status

    if response.status_code in (403, 429):
        return STATUS_BLOCKED, technical_status

    return STATUS_FAILED, technical_status


def read_settings(config_path: str = "settings.ini") -> dict:
    path = Path(config_path)

    if not path.exists():
        return DEFAULT_SETTINGS.copy()

    config = configparser.ConfigParser()
    config.read(path, encoding="utf-8")

    if "general" not in config:
        return DEFAULT_SETTINGS.copy()

    section = config["general"]
    raw_column_index = section.get("column_index", "").strip()

    return {
        "input_file": section.get("input_file", "").strip() or None,
        "all_sheets": section.getboolean("all_sheets", fallback=False),
        "sheet": section.get("sheet", "").strip() or None,
        "column_name": section.get("column_name", "").strip() or None,
        "column_index": int(raw_column_index) if raw_column_index else None,
        "blacklist_file": section.get("blacklist_file", "blacklist.txt").strip() or "blacklist.txt",
        "output_file": section.get("output_file", "").strip() or None,
        "timeout": section.getint("timeout", fallback=8),
        "workers": section.getint("workers", fallback=12),
        "details": section.getboolean("details", fallback=False),
    }


def parse_args():
    parser = argparse.ArgumentParser(description="Link checker for Excel files.")

    parser.add_argument("input_file", nargs="?", default=None, help="Путь к Excel-файлу.")
    parser.add_argument("--config", default="settings.ini", help="Путь к settings.ini.")
    parser.add_argument("--known-names", default="known_names.json", help="Путь к known_names.json.")
    parser.add_argument("--sheet", default=None, help="Название листа.")
    parser.add_argument("--all-sheets", action="store_true", help="Обработать все листы.")
    parser.add_argument("--column-name", default=None, help="Название колонки со ссылками.")
    parser.add_argument("--column-index", type=int, default=None, help="Номер колонки со ссылками.")
    parser.add_argument("--blacklist", default=None, help="Путь к blacklist.txt.")
    parser.add_argument("--output", default=None, help="Имя выходного файла.")
    parser.add_argument("--timeout", type=int, default=None, help="Таймаут запроса.")
    parser.add_argument("--workers", type=int, default=None, help="Количество потоков.")
    parser.add_argument("--details", action="store_true", help="Добавить tech-колонку.")

    return parser.parse_args()


def load_known_names(file_path: str) -> dict:
    path = Path(file_path)

    if not path.exists():
        return DEFAULT_KNOWN_NAMES.copy()

    try:
        with path.open("r", encoding="utf-8") as file:
            data = json.load(file)

        if not isinstance(data, dict):
            return DEFAULT_KNOWN_NAMES.copy()

        known_columns = data.get("known_columns", [])
        file_rules = data.get("file_rules", [])

        if not isinstance(known_columns, list):
            known_columns = []

        if not isinstance(file_rules, list):
            file_rules = []

        cleaned_rules = []
        for rule in file_rules:
            if not isinstance(rule, dict):
                continue

            file_contains = str(rule.get("file_contains", "")).strip()
            column_name = str(rule.get("column_name", "")).strip()

            if file_contains and column_name:
                cleaned_rules.append(
                    {
                        "file_contains": file_contains,
                        "column_name": column_name,
                    }
                )

        cleaned_columns = []
        for column in known_columns:
            value = str(column).strip()
            if value:
                cleaned_columns.append(value)

        return {
            "known_columns": cleaned_columns,
            "file_rules": cleaned_rules,
        }

    except (json.JSONDecodeError, OSError):
        return DEFAULT_KNOWN_NAMES.copy()


def save_known_names(file_path: str, data: dict) -> None:
    path = Path(file_path)

    cleaned_data = {
        "known_columns": [],
        "file_rules": [],
    }

    seen_columns = set()
    for column in data.get("known_columns", []):
        value = str(column).strip()
        if not value:
            continue

        key = value.lower()
        if key in seen_columns:
            continue

        seen_columns.add(key)
        cleaned_data["known_columns"].append(value)

    seen_rules = set()
    for rule in data.get("file_rules", []):
        if not isinstance(rule, dict):
            continue

        file_contains = str(rule.get("file_contains", "")).strip()
        column_name = str(rule.get("column_name", "")).strip()

        if not file_contains or not column_name:
            continue

        key = (file_contains.lower(), column_name.lower())
        if key in seen_rules:
            continue

        seen_rules.add(key)
        cleaned_data["file_rules"].append(
            {
                "file_contains": file_contains,
                "column_name": column_name,
            }
        )

    with path.open("w", encoding="utf-8") as file:
        json.dump(cleaned_data, file, ensure_ascii=False, indent=2)


def merge_settings(args, config_data: dict) -> dict:
    return {
        "input_file": args.input_file if args.input_file is not None else config_data["input_file"],
        "all_sheets": args.all_sheets or config_data["all_sheets"],
        "sheet": args.sheet if args.sheet is not None else config_data["sheet"],
        "column_name": args.column_name if args.column_name is not None else config_data["column_name"],
        "column_index": args.column_index if args.column_index is not None else config_data["column_index"],
        "blacklist_file": args.blacklist if args.blacklist is not None else config_data["blacklist_file"],
        "output_file": args.output if args.output is not None else config_data["output_file"],
        "timeout": args.timeout if args.timeout is not None else config_data["timeout"],
        "workers": args.workers if args.workers is not None else config_data["workers"],
        "details": args.details or config_data["details"],
    }


def find_candidate_files_in_dir(directory: Path) -> list[Path]:
    return sorted(
        [
            file for file in directory.glob("*.xlsx")
            if not file.name.startswith("~$")
        ],
        key=lambda item: item.name.lower(),
    )


def resolve_input_file_from_known_names(
    input_file: Optional[str],
    known_names: dict,
) -> tuple[str, Optional[str], str]:
    if input_file:
        path = Path(input_file)
        if not path.exists():
            raise ValueError(f"Файл '{input_file}' не найден.")
        return str(path), None, "explicit"

    current_dir = Path.cwd()
    excel_files = find_candidate_files_in_dir(current_dir)

    if not excel_files:
        raise ValueError("Excel-файл не указан и в текущей папке не найдено ни одного .xlsx файла.")

    matched_files = []
    for file in excel_files:
        file_name_lower = file.name.lower()

        for rule in known_names.get("file_rules", []):
            pattern = str(rule.get("file_contains", "")).strip().lower()
            if pattern and pattern in file_name_lower:
                matched_files.append((file, rule.get("column_name", "").strip(), pattern))

    unique_matches = {}
    for file, column_name, pattern in matched_files:
        unique_matches[file.name.lower()] = (file, column_name, pattern)

    if len(unique_matches) == 1:
        file, column_name, pattern = next(iter(unique_matches.values()))
        print(f"Файл не указан. По known_names.json найден файл: {file.name}")
        return str(file), column_name or None, "known_rule"

    if len(excel_files) == 1:
        print(f"Файл не указан. Использую единственный найденный файл: {excel_files[0].name}")
        return str(excel_files[0]), None, "single_file"

    file_names = "\n".join(f"- {file.name}" for file in excel_files)
    raise ValueError(
        "Excel-файл не указан, а в папке найдено несколько .xlsx файлов.\n"
        "Укажи файл явно или добавь правило в known_names.json.\n"
        f"{file_names}"
    )


def find_column_by_name(sheet, column_name: str) -> int:
    for cell in sheet[1]:
        value = "" if cell.value is None else str(cell.value).strip()
        if value.lower() == column_name.strip().lower():
            return cell.column

    raise ValueError(f"Колонка '{column_name}' не найдена.")


def validate_column_index(sheet, column_index: int) -> int:
    if column_index < 1:
        raise ValueError("Номер колонки должен быть больше 0.")

    if column_index > sheet.max_column:
        raise ValueError(
            f"В листе только {sheet.max_column} колонок, а передан номер {column_index}."
        )

    return column_index


def find_columns_from_candidates(sheet, candidates: list[str]) -> list[tuple[int, str]]:
    matched_columns = []

    lowered_candidates = {item.strip().lower() for item in candidates if str(item).strip()}

    for cell in sheet[1]:
        value = "" if cell.value is None else str(cell.value).strip()
        lowered = value.lower()

        if lowered in lowered_candidates:
            matched_columns.append((cell.column, value))

    return matched_columns


def find_column_by_known_names(sheet, known_names: dict) -> int:
    known_columns = known_names.get("known_columns", [])
    matched_columns = find_columns_from_candidates(sheet, known_columns)

    if len(matched_columns) == 1:
        column_index, column_name = matched_columns[0]
        print(f"Колонка не указана. По known_names.json использую: {column_name}")
        return column_index

    if len(matched_columns) > 1:
        variants = "\n".join(f"- {name} (колонка {index})" for index, name in matched_columns)
        raise ValueError(
            "Найдено несколько колонок из known_names.json. Укажи column_name или column_index явно.\n"
            f"{variants}"
        )

    raise ValueError("Подходящая колонка в known_names.json не найдена.")


def find_column_by_auto_candidates(sheet) -> int:
    matched_columns = find_columns_from_candidates(sheet, list(AUTO_COLUMN_CANDIDATES))

    if len(matched_columns) == 1:
        column_index, column_name = matched_columns[0]
        print(f"Колонка не указана. По автопоиску использую: {column_name}")
        return column_index

    if len(matched_columns) > 1:
        variants = "\n".join(f"- {name} (колонка {index})" for index, name in matched_columns)
        raise ValueError(
            "Найдено несколько подходящих колонок по автопоиску. Укажи column_name или column_index явно.\n"
            f"{variants}"
        )

    raise ValueError("Колонка не указана и автоматически не найдена.")


def resolve_target_column(
    sheet,
    column_name: Optional[str],
    column_index: Optional[int],
    known_names: dict,
    preferred_column_name: Optional[str] = None,
) -> tuple[int, str, str]:
    if column_name:
        resolved_index = find_column_by_name(sheet, column_name)
        return resolved_index, column_name, "explicit_name"

    if column_index is not None:
        resolved_index = validate_column_index(sheet, column_index)
        header_value = sheet.cell(row=1, column=resolved_index).value
        resolved_name = "" if header_value is None else str(header_value).strip()
        return resolved_index, resolved_name, "explicit_index"

    if preferred_column_name:
        resolved_index = find_column_by_name(sheet, preferred_column_name)
        return resolved_index, preferred_column_name, "known_rule"

    try:
        resolved_index = find_column_by_known_names(sheet, known_names)
        header_value = sheet.cell(row=1, column=resolved_index).value
        resolved_name = "" if header_value is None else str(header_value).strip()
        return resolved_index, resolved_name, "known_column"
    except ValueError:
        pass

    resolved_index = find_column_by_auto_candidates(sheet)
    header_value = sheet.cell(row=1, column=resolved_index).value
    resolved_name = "" if header_value is None else str(header_value).strip()
    return resolved_index, resolved_name, "auto"


def is_service_header(header_value) -> bool:
    if header_value is None:
        return False

    text = str(header_value).replace("\xa0", " ").strip().lower()

    service_prefixes = (
        "link check",
        "tech",
        "результат проверки",
        "технический статус",
    )

    return any(text.startswith(prefix) for prefix in service_prefixes)


def is_meaningful_cell_value(value) -> bool:
    if value is None:
        return False

    text = str(value).replace("\xa0", " ").strip()
    return bool(text)


def get_source_columns(sheet) -> list[int]:
    source_columns = []

    for col in range(1, sheet.max_column + 1):
        header_value = sheet.cell(row=1, column=col).value

        if is_service_header(header_value):
            break

        source_columns.append(col)

    return source_columns


def get_last_data_row(sheet, source_columns: list[int], start_row: int) -> int:
    last_data_row = start_row - 1

    for row in range(start_row, sheet.max_row + 1):
        row_has_data = any(
            is_meaningful_cell_value(sheet.cell(row=row, column=col).value)
            for col in source_columns
        )

        if row_has_data:
            last_data_row = row

    return last_data_row


def create_result_column(sheet, header_name: str) -> int:
    new_col_index = sheet.max_column + 1
    sheet.cell(row=1, column=new_col_index).value = header_name
    return new_col_index


def collect_unique_values(sheet, target_col: int, start_row: int, end_row: int) -> set[str]:
    values = set()

    for row in range(start_row, end_row + 1):
        raw_value = sheet.cell(row=row, column=target_col).value
        normalized = normalize_url(raw_value)
        values.add(normalized)

    return values


def build_cache(
    unique_values: set[str],
    blacklist: set[str],
    timeout: int,
    max_workers: int,
    progress_callback: Optional[Callable[[int, int], None]] = None,
    cancel_event: Optional[threading.Event] = None,
) -> dict[str, tuple[str, str]]:
    cache = {}

    def worker(value: str) -> tuple[str, tuple[str, str]]:
        return value, check_one_url(value, blacklist, timeout)

    with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
        values_iter = iter(unique_values)
        future_to_value = {}
        total = len(unique_values)
        completed = 0

        for _ in range(min(max_workers, total)):
            if cancel_event is not None and cancel_event.is_set():
                break

            try:
                value = next(values_iter)
            except StopIteration:
                break

            future = executor.submit(worker, value)
            future_to_value[future] = value

        while future_to_value:
            done_futures, _ = concurrent.futures.wait(
                future_to_value,
                return_when=concurrent.futures.FIRST_COMPLETED,
            )

            for future in done_futures:
                value, result = future.result()
                future_to_value.pop(future, None)
                cache[value] = result
                completed += 1

                if progress_callback is not None:
                    progress_callback(completed, total)

                if cancel_event is not None and cancel_event.is_set():
                    continue

                try:
                    next_value = next(values_iter)
                except StopIteration:
                    continue

                next_future = executor.submit(worker, next_value)
                future_to_value[next_future] = next_value

            if cancel_event is not None and cancel_event.is_set():
                for future in future_to_value:
                    future.cancel()
                break

    return cache


def paint_status_cell(cell, status: str) -> None:
    if status not in PROBLEM_STATUSES:
        return

    color = STATUS_COLORS.get(status)
    if not color:
        return

    cell.fill = PatternFill(fill_type="solid", fgColor=color)


def process_sheet(
    sheet,
    column_name: Optional[str],
    column_index: Optional[int],
    preferred_column_name: Optional[str],
    known_names: dict,
    blacklist: set[str],
    timeout: int,
    max_workers: int,
    add_details: bool,
    progress_callback: Optional[Callable[[str, int, int], None]] = None,
    cancel_event: Optional[threading.Event] = None,
) -> tuple[Counter, Optional[str], str]:
    target_col, resolved_column_name, resolution_source = resolve_target_column(
        sheet=sheet,
        column_name=column_name,
        column_index=column_index,
        known_names=known_names,
        preferred_column_name=preferred_column_name,
    )

    source_columns = get_source_columns(sheet)
    if not source_columns:
        raise ValueError("Не удалось определить границы исходной таблицы.")

    start_row = 2
    last_row = get_last_data_row(sheet, source_columns, start_row)

    if last_row < start_row:
        print("Лист пустой: данных в исходной таблице нет.")
        return Counter(), resolved_column_name or None, resolution_source

    run_date = get_run_date_label()

    result_col = create_result_column(sheet, f"link check ({run_date})")
    details_col = (
        create_result_column(sheet, f"tech ({run_date})")
        if add_details
        else None
    )

    unique_values = collect_unique_values(sheet, target_col, start_row, last_row)
    print(f"Уникальных значений для проверки: {len(unique_values)}")

    if progress_callback is not None:
        progress_callback(sheet.title, 0, len(unique_values))

    cache = build_cache(
        unique_values=unique_values,
        blacklist=blacklist,
        timeout=timeout,
        max_workers=max_workers,
        progress_callback=(
            None
            if progress_callback is None
            else lambda done, total: progress_callback(sheet.title, done, total)
        ),
        cancel_event=cancel_event,
    )

    summary = Counter()

    for row in range(start_row, last_row + 1):
        if cancel_event is not None and cancel_event.is_set():
            break

        raw_value = sheet.cell(row=row, column=target_col).value
        normalized = normalize_url(raw_value)

        status, tech_status = cache.get(
            normalized,
            (STATUS_FAILED, TECH_REQUEST_ERROR),
        )

        result_cell = sheet.cell(row=row, column=result_col)
        result_cell.value = status
        paint_status_cell(result_cell, status)

        if add_details and details_col is not None:
            sheet.cell(row=row, column=details_col).value = tech_status

        summary[status] += 1

    return summary, resolved_column_name or None, resolution_source


def build_output_path(input_file: str, output_file: Optional[str]) -> str:
    if output_file:
        return output_file

    input_path = Path(input_file)
    return str(input_path.with_name(f"{input_path.stem}_checked{input_path.suffix}"))


def process_workbook(
    input_file: str,
    sheet_name: Optional[str],
    process_all_sheets: bool,
    column_name: Optional[str],
    column_index: Optional[int],
    preferred_column_name: Optional[str],
    known_names: dict,
    blacklist_file: str,
    output_file: Optional[str],
    timeout: int,
    max_workers: int,
    add_details: bool,
    progress_callback: Optional[Callable[[str, int, int], None]] = None,
    cancel_event: Optional[threading.Event] = None,
) -> tuple[str, Counter, int, int, int, Optional[str], Optional[str]]:
    workbook = openpyxl.load_workbook(input_file)

    try:
        blacklist = load_blacklist(blacklist_file)

        if process_all_sheets:
            sheets_to_process = workbook.sheetnames
        else:
            if sheet_name:
                if sheet_name not in workbook.sheetnames:
                    raise ValueError(f"Лист '{sheet_name}' не найден.")
                sheets_to_process = [sheet_name]
            else:
                sheets_to_process = [workbook.active.title]

        total_summary = Counter()
        processed_sheets = 0
        skipped_sheets = 0
        empty_sheets = 0
        remembered_column_name = None
        remembered_source = None

        for current_sheet_name in sheets_to_process:
            if cancel_event is not None and cancel_event.is_set():
                break

            print(f"\nОбработка листа: {current_sheet_name}")
            sheet = workbook[current_sheet_name]

            try:
                sheet_summary, resolved_column_name, resolution_source = process_sheet(
                    sheet=sheet,
                    column_name=column_name,
                    column_index=column_index,
                    preferred_column_name=preferred_column_name,
                    known_names=known_names,
                    blacklist=blacklist,
                    timeout=timeout,
                    max_workers=max_workers,
                    add_details=add_details,
                    progress_callback=progress_callback,
                    cancel_event=cancel_event,
                )

                if remembered_column_name is None and resolved_column_name:
                    remembered_column_name = resolved_column_name
                    remembered_source = resolution_source

                if sheet_summary:
                    total_summary.update(sheet_summary)
                    processed_sheets += 1
                else:
                    empty_sheets += 1

            except ValueError as error:
                print(f"Лист пропущен: {error}")
                skipped_sheets += 1
                continue

        save_path = build_output_path(input_file, output_file)
        workbook.save(save_path)

        return (
            save_path,
            total_summary,
            processed_sheets,
            skipped_sheets,
            empty_sheets,
            remembered_column_name,
            remembered_source,
        )

    finally:
        workbook.close()


def update_known_names_after_success(
    known_names: dict,
    known_names_path: str,
    resolved_input_file: str,
    resolved_column_name: Optional[str],
    explicit_input_file: Optional[str],
    explicit_column_name: Optional[str],
    preferred_column_name: Optional[str],
    resolution_source: Optional[str],
) -> None:
    if not resolved_column_name:
        return

    input_path = Path(resolved_input_file)
    file_stem = input_path.stem.strip()

    if not file_stem:
        return

    normalized_column = resolved_column_name.strip()
    if not normalized_column:
        return

    if normalized_column.lower() not in {item.lower() for item in known_names.get("known_columns", [])}:
        known_names.setdefault("known_columns", []).append(normalized_column)

    should_add_rule = False

    if explicit_input_file and explicit_column_name:
        should_add_rule = True
    elif preferred_column_name and resolution_source == "known_rule":
        should_add_rule = False
    elif resolution_source in {"explicit_name", "explicit_index", "known_rule", "known_column", "auto"}:
        should_add_rule = True

    if should_add_rule:
        existing_rules = known_names.setdefault("file_rules", [])

        rule_exists = any(
            str(rule.get("file_contains", "")).strip().lower() == file_stem.lower()
            and str(rule.get("column_name", "")).strip().lower() == normalized_column.lower()
            for rule in existing_rules
        )

        if not rule_exists:
            existing_rules.append(
                {
                    "file_contains": file_stem,
                    "column_name": normalized_column,
                }
            )

    save_known_names(known_names_path, known_names)


def print_start_info(
    input_file: str,
    all_sheets: bool,
    sheet: Optional[str],
    column_name: Optional[str],
    column_index: Optional[int],
    preferred_column_name: Optional[str],
    details: bool,
) -> None:
    print("=== START ===")
    print(f"file: {Path(input_file).name}")

    if all_sheets:
        print("sheets: all")
    else:
        print(f"sheets: {sheet if sheet else 'active'}")

    if column_name:
        print(f"column: {column_name}")
    elif column_index is not None:
        print(f"column index: {column_index}")
    elif preferred_column_name:
        print(f"column from known names: {preferred_column_name}")
    else:
        print("column: auto")

    print(f"details: {'on' if details else 'off'}")
    print("=============")


def print_summary(
    summary: Counter,
    output_path: str,
    processed_sheets: int,
    skipped_sheets: int,
    empty_sheets: int,
) -> None:
    ordered_statuses = [
        STATUS_OK,
        STATUS_REDIRECT,
        STATUS_BLACKLIST,
        STATUS_EMPTY,
        STATUS_INVALID,
        STATUS_BLOCKED,
        STATUS_FAILED,
    ]

    total = sum(summary.values())

    print("\n=== SUMMARY ===")
    print(f"processed sheets: {processed_sheets}")
    print(f"skipped sheets: {skipped_sheets}")
    print(f"empty sheets: {empty_sheets}")
    print(f"total rows: {total}")

    for status in ordered_statuses:
        print(f"{status}: {summary.get(status, 0)}")

    print(f"saved to: {output_path}")


def detect_preferred_column_from_rules(
    resolved_input_file: str,
    known_names: dict,
) -> Optional[str]:
    file_name_lower = Path(resolved_input_file).name.lower()

    matched_rules = []
    for rule in known_names.get("file_rules", []):
        pattern = str(rule.get("file_contains", "")).strip().lower()
        column_name = str(rule.get("column_name", "")).strip()

        if pattern and column_name and pattern in file_name_lower:
            matched_rules.append(column_name)

    unique_columns = []
    seen = set()

    for column_name in matched_rules:
        key = column_name.lower()
        if key in seen:
            continue
        seen.add(key)
        unique_columns.append(column_name)

    if len(unique_columns) == 1:
        return unique_columns[0]

    return None


def main():
    args = parse_args()
    config_data = read_settings(args.config)
    settings = merge_settings(args, config_data)
    known_names = load_known_names(args.known_names)

    if settings["all_sheets"] and settings["sheet"]:
        print("Ошибка: укажи либо all_sheets, либо sheet.")
        return

    if settings["column_name"] and settings["column_index"] is not None:
        print("Ошибка: укажи либо column_name, либо column_index.")
        return

    try:
        resolved_input_file, known_rule_column_name, _ = resolve_input_file_from_known_names(
            input_file=settings["input_file"],
            known_names=known_names,
        )

        preferred_column_name = None

        if not settings["column_name"] and settings["column_index"] is None:
            preferred_column_name = known_rule_column_name or detect_preferred_column_from_rules(
                resolved_input_file=resolved_input_file,
                known_names=known_names,
            )

        print_start_info(
            input_file=resolved_input_file,
            all_sheets=settings["all_sheets"],
            sheet=settings["sheet"],
            column_name=settings["column_name"],
            column_index=settings["column_index"],
            preferred_column_name=preferred_column_name,
            details=settings["details"],
        )

        (
            output_path,
            summary,
            processed_sheets,
            skipped_sheets,
            empty_sheets,
            remembered_column_name,
            remembered_source,
        ) = process_workbook(
            input_file=resolved_input_file,
            sheet_name=settings["sheet"],
            process_all_sheets=settings["all_sheets"],
            column_name=settings["column_name"],
            column_index=settings["column_index"],
            preferred_column_name=preferred_column_name,
            known_names=known_names,
            blacklist_file=settings["blacklist_file"],
            output_file=settings["output_file"],
            timeout=settings["timeout"],
            max_workers=settings["workers"],
            add_details=settings["details"],
        )

        print_summary(
            summary=summary,
            output_path=output_path,
            processed_sheets=processed_sheets,
            skipped_sheets=skipped_sheets,
            empty_sheets=empty_sheets,
        )

        if processed_sheets > 0:
            update_known_names_after_success(
                known_names=known_names,
                known_names_path=args.known_names,
                resolved_input_file=resolved_input_file,
                resolved_column_name=remembered_column_name,
                explicit_input_file=args.input_file or settings["input_file"],
                explicit_column_name=settings["column_name"],
                preferred_column_name=preferred_column_name,
                resolution_source=remembered_source,
            )

    except Exception as error:
        print(f"Ошибка: {error}")


if __name__ == "__main__":
    main()
